# Project title         :  Pay Book – Billing and Invoice Generator
# Project for           :  IT Fundamentals and Applications
# Project completed by  :  Muhammad Abdullah Elahi, TC-061
# Project submitted to  :  Dr. Amir Zeb, Lab Instructor and Class Teacher


# ────────────────────────────────────────────────
# 1. Import Required Modules
#    (OpenPyXL for Excel creation and styling)
# ────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ────────────────────────────────────────────────
# 2. Function: export_to_excel(values, filename)
#    Generates a formatted Excel project bill summary
# ────────────────────────────────────────────────
def export_to_excel(values, filename):

    # ────────────────────────────────────────────
    # 2.1 Create a new workbook and select active sheet
    # ────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Bill Summary"

    # ────────────────────────────────────────────
    # 2.2 Add main title
    # ────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    ws["A1"] = "📘 PROJECT BILL SUMMARY"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = PatternFill(start_color="183f6b", end_color="183f6b", fill_type="solid")


    # ────────────────────────────────────────────
    # 2.3 General Information Section
    # ────────────────────────────────────────────
    general_info = [
        ["Project Name", values.get("Project Name", "")],
        ["Client Name", values.get("Client Name", "")],
        ["Date", values.get("Date", "")],
        ["Objective", values.get("Objective", "")],
        ["Time", values.get("Time", "")]
    ]

    ws.append([])  # Empty space row
    ws.append(["General Information"])  # Section header
    ws["A3"].font = Font(bold=True, size=12)

    for row in general_info:
        ws.append(row)

    ws.append([])  # Empty spacing row


    # ────────────────────────────────────────────
    # 2.4 Services Provided Section
    # ────────────────────────────────────────────
    ws.append(["Services Provided"])
    ws["A10"].font = Font(bold=True, size=12)

    headers = ["#", "Service/ Product", "Amount (PKR)", "Discount (%)", "Total (PKR)"]
    ws.append(headers)

    # Define cell borders
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Populate services
    if "Services" in values:
        for idx, s in enumerate(values["Services"], 1):
            ws.append([
                idx,
                s["Service/ Product"],
                s["Amount (PKR)"],
                s["Discount (%)"],
                s["Total (PKR)"],
            ])

        # Apply borders to last row of services
        for col in range(1, 6):
            ws.cell(row=ws.max_row, column=col).border = border


    # ────────────────────────────────────────────
    # 2.5 Auto-fit column widths
    # ────────────────────────────────────────────
    for i, col in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(i)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2


    # ────────────────────────────────────────────
    # 2.6 Billing Summary Section
    # ────────────────────────────────────────────
    ws.append([])  # Empty row
    ws.append(["Billing Summary"])
    ws["A{}".format(ws.max_row)].font = Font(bold=True, size=12)

    ws.append(["Total Amount (PKR)", values.get("Total Amount (PKR)", 0)])
    ws.append(["Advance (50%)", values.get("Advance", 0)])
    ws.append(["Remaining Balance", values.get("Total Amount (PKR)", 0) - values.get("Advance", 0)])


    # ────────────────────────────────────────────
    # 2.7 Save Workbook
    # ────────────────────────────────────────────
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    wb.save(filename)
    print(f"✅ Excel file exported successfully: {filename}")
